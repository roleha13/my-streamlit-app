import pandas as pd
from datetime import datetime
from io import BytesIO
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font

# ================= CORE FUNCTIONS =================

def corrected_month_to_period(dt):
    if pd.isna(dt):
        return ""
    month = dt.month
    year = dt.year + 1 if month >= 7 else dt.year
    period_num = ((month - 7) % 12) + 2
    return f"{year}/{period_num:03d}"

def format_month_label(dt):
    if pd.isna(dt):
        return ""
    return dt.strftime("%B'%y")

def process_file(uploaded_file):
    xls = pd.ExcelFile(uploaded_file)
    rows = []

    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)

        if df.shape[1] < 16:
            continue

        mask = df.iloc[:, 6].str.contains("POS Goldenkey Ltd", case=False, na=False)
        filtered = df[mask].copy()

        if filtered.empty:
            continue

        selected = filtered.iloc[:, [1, 9, 15]].copy()
        selected.columns = ["ColB", "ColJ", "ColP"]

        for _, r in selected.iterrows():
            dt = pd.to_datetime(r["ColB"], errors="coerce")
            if pd.isna(dt):
                continue

            rows.append({
                "1": "1;3;6",
                "TRANSACTION REFERENCE": format_month_label(dt),
                "DESCRIPTION": r["ColP"] if pd.notna(r["ColP"]) else "",
                "ACCOUNT CODE": "440100",
                "TRANSACTION DATE": dt,
                "PERIOD": corrected_month_to_period(dt),
                "BASE AMOUNT": float(str(r["ColJ"]).replace(",", "").strip() or 0),
                "DEBIT/CREDIT": "D",
                "TRANSACTION CURRENCY": "KSH"
            })

    return pd.DataFrame(rows)

def generate_master(files):
    all_data = []

    for file in files:
        df = process_file(file)
        if not df.empty:
            all_data.append(df)

    if not all_data:
        return None

    df = pd.concat(all_data, ignore_index=True)
    df["TRANSACTION DATE"] = pd.to_datetime(df["TRANSACTION DATE"])

    # Sort by date and description
    df = df.sort_values(["TRANSACTION DATE", "DESCRIPTION"])
    
    # Write initial Excel
    output = BytesIO()
    df.to_excel(output, index=False, engine="openpyxl")
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    # ========== Insert FOOD INV formula rows ==========
    current_row = 2  # Excel row 1 = header

    for date, group in df.groupby("TRANSACTION DATE", sort=True):
        n = len(group)
        start_row = current_row
        end_row = current_row + n - 1

        month_label = format_month_label(date)  # <<< ADD THIS LINE
        
        # Insert a row after the last transaction of this group
        ws.insert_rows(end_row + 1)                        
        
        # 1 Column A (1)
        ws.cell(row=end_row + 1, column=1, value="1;3;6")
        # TRANSACTION REFERENCE Column B (2)
        ws.cell(row=end_row + 1, column=2, value=month_label)
        # DESCRIPTION column C (3)
        ws.cell(row=end_row + 1, column=3, value=f"FOOD INV {date.strftime('%d.%m.%Y')}")
        # ACCOUNT CODE column D (4)
        ws.cell(row=end_row + 1, column=4, value="CT00311")
        # TRANSACTION DATE column E (5)
        ws.cell(row=end_row + 1, column=5, value=date)
        # PERIOD column F (6)
        ws.cell(row=end_row + 1, column=6, value=corrected_month_to_period(date))
        # BASE AMOUNT column G (7) — formula
        ws.cell(row=end_row + 1, column=7, value=f"=SUM(G{start_row}:G{end_row})")
        # DEBIT/CREDIT column H (8)
        ws.cell(row=end_row + 1, column=8, value="C")
        # TRANSACTION CURRENCY column I (9)
        ws.cell(row=end_row + 1, column=9, value="KSH")

        # Bold FOOD INV row
        for cell in ws[end_row + 1]:
            cell.font = Font(bold=True)

        # Update current_row for next group
        current_row = end_row + 2

    # ========== Format Excel ==========
    # Format BASE AMOUNT (Column G)
    for cell in ws["G"][1:]:
        cell.number_format = '#,##0.00'

    # Format TRANSACTION DATE (Column E)
    for cell in ws["E"][1:]:
        cell.number_format = 'dd/mm/yyyy'
        cell.alignment = cell.alignment.copy(horizontal='right')

    # Save to BytesIO
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return final_output

# ================= STREAMLIT UI =================

st.set_page_config(page_title="GKC Food Invoice Generator", layout="centered")
st.title("📊 GKC CRJ Data Extraction & Upload Preparation Tool (Web Version)")
st.markdown("Upload your CRJ Excel files and generate a formatted master file with live SUM formulas.")

uploaded_files = st.file_uploader(
    "📂 Upload Excel Files",
    type=["xlsx", "xls"],
    accept_multiple_files=True
)

if st.button("▶️ Generate Master Excel"):
    if uploaded_files:
        with st.spinner("Processing files..."):
            result = generate_master(uploaded_files)

        if result:
            st.success("✅ File generated successfully!")
            st.download_button(
                label="📥 Download Excel",
                data=result,
                file_name="CRJ_ALL_TRANSFORMED.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("⚠️ No matching data found in uploaded files.")
    else:
        st.error("❌ Please upload at least one Excel file.")
